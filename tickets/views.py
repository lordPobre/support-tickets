from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q
from django.contrib import messages
from django.utils import timezone
from .models import Company, Ticket, TicketStatus, TicketPriority
from .forms import TicketPublicForm, TicketCommentForm, TicketSearchForm
from django.http import HttpResponse
from .exports import generate_ticket_pdf, generate_ticket_image


# ─────────────────────────────────────────────────────────────
#  INTERNAL DASHBOARD (requires login)
# ─────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    tickets = Ticket.objects.select_related("company", "status", "priority")
    total = tickets.count()
    open_tickets = tickets.exclude(status__is_closed_state=True).count()
    billed_total = tickets.filter(is_billed=True).aggregate(s=Sum("assigned_value"))["s"] or 0
    unbilled_total = (
        tickets.filter(is_billed=False)
        .exclude(assigned_value__isnull=True)
        .aggregate(s=Sum("assigned_value"))["s"] or 0
    )
    recent_tickets = tickets.order_by("-created_at")[:10]
    all_tickets = tickets.order_by("-created_at").select_related("company", "status", "priority")
    statuses = TicketStatus.objects.annotate(count=Count("tickets"))
    companies_stats = (
        Company.objects.filter(is_active=True)
        .annotate(
            ticket_count=Count("tickets"),
            open_count=Count("tickets", filter=Q(tickets__status__is_closed_state=False)),
            pending_value=Sum("tickets__assigned_value", filter=Q(tickets__is_billed=False)),
        )
        .order_by("-ticket_count")[:8]
    )
    context = {
        "total": total,
        "open_tickets": open_tickets,
        "billed_total": billed_total,
        "unbilled_total": unbilled_total,
        "recent_tickets": recent_tickets,
        "all_tickets": all_tickets,
        "statuses": statuses,
        "companies_stats": companies_stats,
    }
    return render(request, "tickets/dashboard.html", context)


@login_required
def ticket_list(request):
    tickets = Ticket.objects.select_related("company", "status", "priority").order_by("-created_at")
    company_id  = request.GET.get("company")
    status_id   = request.GET.get("status")
    priority_id = request.GET.get("priority")
    category    = request.GET.get("category")
    billed      = request.GET.get("billed")
    date_from   = request.GET.get("date_from", "")
    date_to     = request.GET.get("date_to", "")
    q           = request.GET.get("q", "").strip()

    if company_id:
        tickets = tickets.filter(company_id=company_id)
    if status_id:
        tickets = tickets.filter(status_id=status_id)
    if priority_id:
        tickets = tickets.filter(priority_id=priority_id)
    if category:
        tickets = tickets.filter(category=category)
    if billed == "1":
        tickets = tickets.filter(is_billed=True)
    elif billed == "0":
        tickets = tickets.filter(is_billed=False)
    if date_from:
        tickets = tickets.filter(created_at__date__gte=date_from)
    if date_to:
        tickets = tickets.filter(created_at__date__lte=date_to)
    if q:
        tickets = tickets.filter(
            Q(token__icontains=q) | Q(subject__icontains=q) |
            Q(requester_name__icontains=q) | Q(requester_email__icontains=q)
        )

    context = {
        "tickets": tickets,
        "companies": Company.objects.filter(is_active=True),
        "statuses": TicketStatus.objects.all(),
        "priorities": TicketPriority.objects.all(),
        "category_choices": Ticket.CATEGORY_CHOICES,
        "filters": {
            "company": company_id, "status": status_id, "priority": priority_id,
            "category": category, "billed": billed,
            "date_from": date_from, "date_to": date_to, "q": q,
        }
    }
    return render(request, "tickets/ticket_list.html", context)


@login_required
def ticket_detail(request, token):
    ticket = get_object_or_404(Ticket, token=token)
    comment_form = TicketCommentForm()

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "comment":
            comment_form = TicketCommentForm(request.POST)
            if comment_form.is_valid():
                comment = comment_form.save(commit=False)
                comment.ticket = ticket
                comment.author_name = request.user.get_full_name() or request.user.username
                comment.author_email = request.user.email
                comment.is_staff = True
                comment.is_internal = request.POST.get("is_internal") == "on"
                comment.save()
                if not comment.is_internal:
                    ticket.send_status_update_email("Nueva respuesta de soporte")
                messages.success(request, "Comentario agregado.")
                return redirect("ticket_detail", token=token)

        elif action == "update_value":
            value = request.POST.get("assigned_value", "").strip()
            try:
                ticket.assigned_value = float(value) if value else None
                ticket.save(update_fields=["assigned_value"])
                messages.success(request, "Valor actualizado.")
            except ValueError:
                messages.error(request, "Valor inválido.")
            return redirect("ticket_detail", token=token)

        elif action == "update_status":
            status_id = request.POST.get("status_id")
            try:
                status = TicketStatus.objects.get(pk=status_id)
                old_status = ticket.status
                ticket.status = status
                if status.is_closed_state and not ticket.closed_at:
                    ticket.closed_at = timezone.now()
                ticket.save(update_fields=["status", "closed_at"])
                if old_status != status:
                    ticket.send_status_update_email(status.name)
                messages.success(request, f"Estado cambiado a {status.name}.")
            except TicketStatus.DoesNotExist:
                messages.error(request, "Estado no encontrado.")
            return redirect("ticket_detail", token=token)

        elif action == "toggle_billed":
            ticket.is_billed = not ticket.is_billed
            ticket.save(update_fields=["is_billed"])
            messages.success(request, "Estado de facturación actualizado.")
            return redirect("ticket_detail", token=token)

    context = {
        "ticket": ticket,
        "comment_form": comment_form,
        "statuses": TicketStatus.objects.all(),
        "priorities": TicketPriority.objects.all(),
    }
    return render(request, "tickets/ticket_detail.html", context)


# ─────────────────────────────────────────────────────────────
#  PUBLIC PORTAL (no login required)
# ─────────────────────────────────────────────────────────────

CATEGORY_META = {
    "software": {
        "icon": "💻", "color": "#6366F1", "bg": "#EEF2FF",
        "border": "#A5B4FC", "label": "Software",
        "desc": "Problemas con aplicaciones, sistemas operativos o software",
    },
    "hardware": {
        "icon": "🖥️", "color": "#D97706", "bg": "#FFFBEB",
        "border": "#FCD34D", "label": "Hardware",
        "desc": "Fallas en equipos, periféricos o componentes físicos",
    },
    "email": {
        "icon": "✉️", "color": "#059669", "bg": "#ECFDF5",
        "border": "#6EE7B7", "label": "Email / Correo",
        "desc": "Configuración, acceso o problemas con correo electrónico",
    },
}


def portal_home(request, company_slug):
    company = get_object_or_404(Company, slug=company_slug, is_active=True)
    has_registered_users = company.users.filter(is_active=True).exists()
    search_form = TicketSearchForm()
    create_form = TicketPublicForm(company=company)

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "search":
            search_form = TicketSearchForm(request.POST)
            if search_form.is_valid():
                token_val = search_form.cleaned_data["token"].strip().upper()
                try:
                    ticket = Ticket.objects.get(token=token_val, company=company)
                    return redirect("portal_ticket", company_slug=company_slug, token=ticket.token)
                except Ticket.DoesNotExist:
                    messages.error(request, "No se encontró un ticket con ese código.")

        elif action == "create":
            create_form = TicketPublicForm(company=company, data=request.POST, files=request.FILES)
            if create_form.is_valid():
                default_status = TicketStatus.objects.filter(is_closed_state=False).order_by("order").first()
                ticket = create_form.save(commit=False)
                ticket.company = company
                ticket.status = default_status
                ticket.save()

                for f in request.FILES.getlist("attachments"):
                    from .models import TicketAttachment
                    TicketAttachment.objects.create(ticket=ticket, file=f, original_name=f.name)

                ticket.send_confirmation_email()
                messages.success(
                    request,
                    f"¡Ticket creado! Tu código es <strong>{ticket.token}</strong>. "
                    f"Recibirás un email de confirmación."
                )
                return redirect("portal_ticket", company_slug=company_slug, token=ticket.token)

    # Historial de tickets de la empresa
    history_qs = company.tickets.select_related("status", "priority").order_by("-created_at")
    filter_category = request.GET.get("cat", "")
    filter_status   = request.GET.get("st", "")
    if filter_category:
        history_qs = history_qs.filter(category=filter_category)
    if filter_status:
        history_qs = history_qs.filter(status__slug=filter_status)
    company_statuses = TicketStatus.objects.filter(tickets__company=company).distinct()

    context = {
        "company": company,
        "has_registered_users": has_registered_users,
        "search_form": search_form,
        "create_form": create_form,
        "priorities": TicketPriority.objects.all(),
        "category_meta": CATEGORY_META,
        "history_tickets": history_qs,
        "company_statuses": company_statuses,
        "filter_category": filter_category,
        "filter_status": filter_status,
    }
    return render(request, "tickets/portal_home.html", context)


def portal_ticket(request, company_slug, token):
    company = get_object_or_404(Company, slug=company_slug, is_active=True)
    ticket = get_object_or_404(Ticket, token=token, company=company)
    comment_form = TicketCommentForm()

    if request.method == "POST":
        comment_form = TicketCommentForm(request.POST)
        if comment_form.is_valid():
            comment = comment_form.save(commit=False)
            comment.ticket = ticket
            comment.author_name = ticket.requester_name
            comment.author_email = ticket.requester_email
            comment.is_staff = False
            comment.is_internal = False
            comment.save()
            messages.success(request, "Tu mensaje fue enviado.")
            return redirect("portal_ticket", company_slug=company_slug, token=token)

    public_comments = ticket.comments.filter(is_internal=False).order_by("created_at")
    context = {
        "company": company,
        "ticket": ticket,
        "comment_form": comment_form,
        "public_comments": public_comments,
        "category_meta": CATEGORY_META,
    }
    return render(request, "tickets/portal_ticket.html", context)


# ─────────────────────────────────────────────────────────────
#  EXPORT: PDF & IMAGE
# ─────────────────────────────────────────────────────────────

@login_required
def ticket_pdf(request, token):
    ticket = get_object_or_404(Ticket, token=token)
    buf = generate_ticket_pdf(ticket)
    response = HttpResponse(buf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="ticket-{ticket.token}.pdf"'
    return response


@login_required
def ticket_image(request, token):
    ticket = get_object_or_404(Ticket, token=token)
    buf = generate_ticket_image(ticket)
    response = HttpResponse(buf, content_type="image/png")
    response["Content-Disposition"] = f'attachment; filename="ticket-{ticket.token}.png"'
    return response


def portal_ticket_pdf(request, company_slug, token):
    company = get_object_or_404(Company, slug=company_slug, is_active=True)
    ticket = get_object_or_404(Ticket, token=token, company=company)
    buf = generate_ticket_pdf(ticket)
    response = HttpResponse(buf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="ticket-{ticket.token}.pdf"'
    return response


def portal_ticket_image(request, company_slug, token):
    company = get_object_or_404(Company, slug=company_slug, is_active=True)
    ticket = get_object_or_404(Ticket, token=token, company=company)
    buf = generate_ticket_image(ticket)
    response = HttpResponse(buf, content_type="image/png")
    response["Content-Disposition"] = f'attachment; filename="ticket-{ticket.token}.png"'
    return response


# ─────────────────────────────────────────────────────────────
#  EXPORT EXCEL
# ─────────────────────────────────────────────────────────────

@login_required
def ticket_export_excel(request):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    # ── Aplicar los mismos filtros que ticket_list ──
    tickets = Ticket.objects.select_related("company", "status", "priority").order_by("-created_at")

    company_id  = request.GET.get("company")
    status_id   = request.GET.get("status")
    priority_id = request.GET.get("priority")
    category    = request.GET.get("category")
    billed      = request.GET.get("billed")
    date_from   = request.GET.get("date_from", "")
    date_to     = request.GET.get("date_to", "")
    q           = request.GET.get("q", "").strip()

    if company_id:
        tickets = tickets.filter(company_id=company_id)
    if status_id:
        tickets = tickets.filter(status_id=status_id)
    if priority_id:
        tickets = tickets.filter(priority_id=priority_id)
    if category:
        tickets = tickets.filter(category=category)
    if billed == "1":
        tickets = tickets.filter(is_billed=True)
    elif billed == "0":
        tickets = tickets.filter(is_billed=False)
    if date_from:
        tickets = tickets.filter(created_at__date__gte=date_from)
    if date_to:
        tickets = tickets.filter(created_at__date__lte=date_to)
    if q:
        tickets = tickets.filter(
            Q(token__icontains=q) | Q(subject__icontains=q) |
            Q(requester_name__icontains=q) | Q(requester_email__icontains=q)
        )

    # ── Construir Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # Colores
    INDIGO      = "4F46E5"
    INDIGO_LIGHT= "EEF2FF"
    SLATE_100   = "F1F5F9"
    SLATE_200   = "E2E8F0"
    WHITE       = "FFFFFF"
    EMERALD     = "059669"
    SLATE_700   = "334155"

    header_font  = Font(name="Arial", bold=True, color=WHITE, size=10)
    header_fill  = PatternFill("solid", fgColor=INDIGO)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin", color=SLATE_200),
        right=Side(style="thin", color=SLATE_200),
        top=Side(style="thin", color=SLATE_200),
        bottom=Side(style="thin", color=SLATE_200),
    )
    cell_align   = Alignment(vertical="center", wrap_text=True)
    alt_fill     = PatternFill("solid", fgColor=SLATE_100)

    # ── Título ──
    title_label = "Reporte de Tickets de Soporte"
    if date_from and date_to:
        title_label += f" — {date_from} al {date_to}"
    elif date_from:
        title_label += f" — desde {date_from}"
    elif date_to:
        title_label += f" — hasta {date_to}"

    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = title_label
    title_cell.font  = Font(name="Arial", bold=True, color=WHITE, size=13)
    title_cell.fill  = PatternFill("solid", fgColor=INDIGO)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:L2")
    ws["A2"].value = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')}  |  {tickets.count()} tickets"
    ws["A2"].font  = Font(name="Arial", color=SLATE_700, size=9, italic=True)
    ws["A2"].fill  = PatternFill("solid", fgColor=INDIGO_LIGHT)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Encabezados ──
    headers = [
        "Token", "Empresa", "Solicitante", "Email",
        "Categoría", "Estado", "Prioridad",
        "Asunto", "Valor (CLP)", "Facturado",
        "Fecha Creación", "Fecha Cierre"
    ]
    col_widths = [15, 20, 22, 28, 12, 16, 12, 35, 14, 11, 18, 18]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22

    # ── Datos ──
    CATEGORY_LABELS = {"software": "💻 Software", "hardware": "🖥️ Hardware", "email": "✉️ Email"}

    for row_idx, ticket in enumerate(tickets, 4):
        is_alt = (row_idx % 2 == 0)
        row_fill = PatternFill("solid", fgColor=SLATE_100) if is_alt else None

        values = [
            ticket.token,
            ticket.company.name,
            ticket.requester_name,
            ticket.requester_email,
            CATEGORY_LABELS.get(ticket.category, ticket.category),
            ticket.status.name if ticket.status else "—",
            ticket.priority.name if ticket.priority else "—",
            ticket.subject,
            int(ticket.assigned_value) if ticket.assigned_value is not None else "",
            "Sí" if ticket.is_billed else "No",
            ticket.created_at.strftime("%d/%m/%Y %H:%M"),
            ticket.closed_at.strftime("%d/%m/%Y %H:%M") if ticket.closed_at else "—",
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = cell_align
            cell.border    = thin_border
            if row_fill:
                cell.fill = row_fill

            # Formato moneda para valor
            if col_idx == 9 and isinstance(value, int):
                cell.number_format = '#,##0'
                cell.font = Font(name="Arial", size=9, color=EMERALD, bold=True)

            # Color facturado
            if col_idx == 10:
                cell.font = Font(
                    name="Arial", size=9,
                    color=EMERALD if value == "Sí" else "D97706",
                    bold=True
                )

        ws.row_dimensions[row_idx].height = 18

    # ── Fila de totales ──
    last_data_row = 3 + tickets.count()
    total_row = last_data_row + 1

    ws.merge_cells(f"A{total_row}:H{total_row}")
    total_label = ws[f"A{total_row}"]
    total_label.value = f"TOTAL — {tickets.count()} tickets"
    total_label.font  = Font(name="Arial", bold=True, size=10, color=WHITE)
    total_label.fill  = PatternFill("solid", fgColor=INDIGO)
    total_label.alignment = Alignment(horizontal="right", vertical="center")

    total_value = ws.cell(row=total_row, column=9)
    total_value.value  = f"=SUM(I4:I{last_data_row})"
    total_value.font   = Font(name="Arial", bold=True, size=10, color=WHITE)
    total_value.fill   = PatternFill("solid", fgColor=INDIGO)
    total_value.number_format = '#,##0'
    total_value.alignment = Alignment(horizontal="right", vertical="center")

    for col_idx in range(10, 13):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = PatternFill("solid", fgColor=INDIGO)

    ws.row_dimensions[total_row].height = 22

    # ── Freeze panes y filtros ──
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:L{last_data_row}"

    # ── Respuesta HTTP ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if date_from or date_to:
        rango = f"{date_from or 'inicio'}_{date_to or 'hoy'}"
        filename = f"tickets_{rango}.xlsx"
    else:
        filename = "tickets_todos.xlsx"
    response = HttpResponse(
        buf,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────
#  INVENTARIO
# ─────────────────────────────────────────────────────────────

@login_required
def inventory_list(request):
    from .models import Equipment
    equipments = Equipment.objects.select_related("company", "assigned_to").order_by("company", "device_type", "brand")

    company_id  = request.GET.get("company", "")
    status      = request.GET.get("status", "")
    device_type = request.GET.get("device_type", "")
    q           = request.GET.get("q", "").strip()

    if company_id:
        equipments = equipments.filter(company_id=company_id)
    if status:
        equipments = equipments.filter(status=status)
    if device_type:
        equipments = equipments.filter(device_type=device_type)
    if q:
        equipments = equipments.filter(
            Q(brand__icontains=q) | Q(model__icontains=q) |
            Q(serial_number__icontains=q) | Q(assigned_to__name__icontains=q)
        )

    from .models import Equipment as Eq
    context = {
        "equipments": equipments,
        "companies": Company.objects.filter(is_active=True),
        "status_choices": Eq.STATUS_CHOICES,
        "device_type_choices": Eq.DEVICE_TYPES,
        "filters": {"company": company_id, "status": status, "device_type": device_type, "q": q},
        "total": equipments.count(),
        "active_count": equipments.filter(status="active").count(),
        "maintenance_count": equipments.filter(status="maintenance").count(),
    }
    return render(request, "tickets/inventory_list.html", context)


@login_required
def inventory_detail(request, pk):
    from .models import Equipment
    equipment = get_object_or_404(Equipment, pk=pk)
    # Tickets relacionados al usuario asignado en la misma empresa
    related_tickets = []
    if equipment.assigned_to:
        related_tickets = Ticket.objects.filter(
            company=equipment.company,
            company_user=equipment.assigned_to
        ).order_by("-created_at")[:5]

    context = {"equipment": equipment, "related_tickets": related_tickets}
    return render(request, "tickets/inventory_detail.html", context)


@login_required
def inventory_add(request):
    """Create a new equipment."""
    from .forms import EquipmentForm
    from .models import Equipment, CompanyUser

    form = EquipmentForm(request.POST or None, request.FILES or None)

    # AJAX: filter assigned_to by company
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        company_id = request.GET.get("company_id")
        users = CompanyUser.objects.filter(company_id=company_id, is_active=True).values("id", "name")
        from django.http import JsonResponse
        return JsonResponse({"users": list(users)})

    if request.method == "POST" and form.is_valid():
        equipment = form.save()
        messages.success(request, f"Equipo {equipment.brand} {equipment.model} registrado correctamente.")
        return redirect("inventory_detail", pk=equipment.pk)

    context = {"form": form, "title": "Agregar equipo", "is_edit": False}
    return render(request, "tickets/inventory_form.html", context)


@login_required
def inventory_edit(request, pk):
    """Edit existing equipment."""
    from .forms import EquipmentForm
    from .models import Equipment, CompanyUser

    equipment = get_object_or_404(Equipment, pk=pk)

    # AJAX: filter assigned_to by company
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        company_id = request.GET.get("company_id")
        users = CompanyUser.objects.filter(company_id=company_id, is_active=True).values("id", "name")
        from django.http import JsonResponse
        return JsonResponse({"users": list(users)})

    form = EquipmentForm(request.POST or None, request.FILES or None, instance=equipment)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Equipo actualizado correctamente.")
        return redirect("inventory_detail", pk=equipment.pk)

    context = {"form": form, "equipment": equipment, "title": "Editar equipo", "is_edit": True}
    return render(request, "tickets/inventory_form.html", context)


@login_required
def inventory_delete(request, pk):
    """Delete equipment via POST."""
    from .models import Equipment
    equipment = get_object_or_404(Equipment, pk=pk)
    if request.method == "POST":
        name = f"{equipment.brand} {equipment.model}"
        equipment.delete()
        messages.success(request, f"Equipo {name} eliminado.")
        return redirect("inventory_list")
    return redirect("inventory_detail", pk=pk)


# ─────────────────────────────────────────────────────────────
#  PORTAL LOGIN / INVENTARIO
# ─────────────────────────────────────────────────────────────

def _portal_session_key(company_slug):
    return f"portal_user_{company_slug}"


def portal_login(request, company_slug):
    company = get_object_or_404(Company, slug=company_slug, is_active=True)
    error = None

    if request.method == "POST":
        email = request.POST.get("email", "").strip().lower()
        from .models import CompanyUser
        try:
            user = CompanyUser.objects.get(
                company=company, email__iexact=email, is_active=True
            )
            if not user.is_manager:
                error = "Tu usuario no tiene acceso al inventario. Contacta a tu administrador."
            else:
                request.session[_portal_session_key(company_slug)] = {
                    "user_id": user.pk,
                    "user_name": user.name,
                    "user_email": user.email,
                }
                next_url = request.GET.get("next", "")
                if next_url:
                    return redirect(next_url)
                return redirect("portal_inventory", company_slug=company_slug)
        except CompanyUser.DoesNotExist:
            error = "Email no encontrado o no autorizado para esta empresa."

    context = {"company": company, "error": error}
    return render(request, "tickets/portal_login.html", context)


def portal_logout(request, company_slug):
    request.session.pop(_portal_session_key(company_slug), None)
    return redirect("portal_home", company_slug=company_slug)


def portal_inventory(request, company_slug):
    company = get_object_or_404(Company, slug=company_slug, is_active=True)

    # Check session
    session_data = request.session.get(_portal_session_key(company_slug))
    if not session_data:
        from django.urls import reverse
        login_url = reverse("portal_login", kwargs={"company_slug": company_slug})
        return redirect(f"{login_url}?next=/portal/{company_slug}/inventario/")

    from .models import Equipment, CompanyUser
    try:
        portal_user = CompanyUser.objects.get(pk=session_data["user_id"], is_active=True)
    except CompanyUser.DoesNotExist:
        request.session.pop(_portal_session_key(company_slug), None)
        return redirect("portal_login", company_slug=company_slug)

    # All company equipment
    equipments = Equipment.objects.filter(company=company).select_related("assigned_to").order_by("device_type", "brand")

    context = {
        "company": company,
        "portal_user": portal_user,
        "equipments": equipments,
        "my_equipment": equipments.filter(assigned_to=portal_user),
    }
    return render(request, "tickets/portal_inventory.html", context)
